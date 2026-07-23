"""Excel (.xlsx) file extractor."""

import os
from typing import List, Optional

from models.rfq_data import RFQData, RFQItem
from extractors.base import BaseExtractor


class XLSXExtractor(BaseExtractor):
    """Extracts RFQ data from Excel spreadsheets.

    Strategy:
      - Auto-detect header row (looks for 'description', 'qty', etc.)
      - Map columns to RFQ fields
      - Read all data rows
    """

    # Column name patterns (case-insensitive matching)
    DESC_PATTERNS = ("description", "item description", "material", "product", "scope")
    QTY_PATTERNS = ("qty", "quantity", "qnty", "amount", "الكمية")
    MFG_PATTERNS = ("manufacturer", "mfg", "brand", "make", "المصنع")
    MODEL_PATTERNS = ("model", "model number", "model no", "الموديل")
    PARTNO_PATTERNS = ("part number", "part no", "p/n", "pn", "رقم القطعة")
    ITEM_PATTERNS = ("item", "no", "s.no", "sr", "#", "الرقم")

    @staticmethod
    def can_handle(file_path: str) -> bool:
        ext = file_path.lower()
        return ext.endswith(".xlsx") or ext.endswith(".xls")

    def extract(self, file_path: str) -> RFQData:
        import openpyxl

        data = RFQData(source_file_path=file_path)

        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            data.raw_extracted_text = f"[ERROR] Could not open Excel file: {e}"
            return data

        ws = wb.active
        if ws is None:
            data.raw_extracted_text = "[ERROR] No active worksheet found."
            return data

        # Find header row and column mapping
        col_map, header_row = self._find_columns(ws)
        if not col_map:
            # No clear header found — dump all content as raw text
            data.raw_extracted_text = self._dump_all_cells(ws)
            return data

        # Read data rows
        items: List[RFQItem] = []
        raw_lines = [f"[Header at row {header_row}] {col_map}"]
        idx = 1

        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            cells = {col: row[col].value for col in col_map if col < len(row)}

            desc = str(cells.get(col_map.get("description", -1), "") or "").strip()
            qty = str(cells.get(col_map.get("quantity", -1), "") or "").strip()

            if not desc and not qty:
                continue  # Skip empty rows

            item = RFQItem(
                index=idx,
                description=desc,
                quantity=qty,
                manufacturer=str(
                    cells.get(col_map.get("manufacturer", -1), "") or ""
                ).strip() or None,
                model=str(
                    cells.get(col_map.get("model", -1), "") or ""
                ).strip() or None,
                part_number=str(
                    cells.get(col_map.get("part_number", -1), "") or ""
                ).strip() or None,
            )
            items.append(item)
            raw_lines.append(f"Row {idx}: {desc} | Qty: {qty}")
            idx += 1

        data.items = items
        data.raw_extracted_text = "\n".join(raw_lines)

        # Try to get subject from filename
        data.tender_subject = os.path.splitext(os.path.basename(file_path))[0]

        wb.close()
        return data

    def _find_columns(self, ws) -> tuple:
        """Find the header row and map columns to fields."""
        for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=False), 1):
            cell_texts = []
            for cell in row:
                val = str(cell.value or "").strip().lower()
                cell_texts.append(val)

            # Check if this row looks like a header
            has_desc = any(
                any(p in ct for p in self.DESC_PATTERNS) for ct in cell_texts
            )
            has_qty = any(
                any(p in ct for p in self.QTY_PATTERNS) for ct in cell_texts
            )

            if has_desc or has_qty:
                col_map = {}
                for ci, ct in enumerate(cell_texts):
                    if any(p in ct for p in self.DESC_PATTERNS):
                        col_map["description"] = ci
                    elif any(p in ct for p in self.QTY_PATTERNS):
                        col_map["quantity"] = ci
                    elif any(p in ct for p in self.MFG_PATTERNS):
                        col_map["manufacturer"] = ci
                    elif any(p in ct for p in self.MODEL_PATTERNS):
                        col_map["model"] = ci
                    elif any(p in ct for p in self.PARTNO_PATTERNS):
                        col_map["part_number"] = ci
                    elif any(p in ct for p in self.ITEM_PATTERNS):
                        col_map["item"] = ci
                return col_map, row_idx

        return {}, 0

    def _dump_all_cells(self, ws) -> str:
        """Dump all non-empty cells as raw text for manual review."""
        lines = []
        for row in ws.iter_rows(max_row=100, values_only=True):
            vals = [str(c) for c in row if c is not None]
            if vals:
                lines.append(" | ".join(vals))
        return "\n".join(lines)
